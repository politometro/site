import os
import re
import openpyxl

workspace = "."
excel_path = os.path.join(workspace, "Polítólogo Português.xlsx")
data_dir = os.path.join(workspace, "data")

if not os.path.exists(excel_path):
    print(f"Error: Excel file not found at {excel_path}")
    exit(1)

# Load workbook
wb = openpyxl.load_workbook(excel_path)
sheet = wb["Folha1"]

# Let's inspect rows and columns to find indices
# Row indices are 1-based in openpyxl
max_row = sheet.max_row
max_col = sheet.max_column

# Get column names
headers = {}
for col in range(1, max_col + 1):
    val = sheet.cell(row=2, column=col).value  # Header seems to be row 2 or 1. Let's find where the columns are.
    # Wait, let's print row 1 and row 2 to see which row has the column names.
    # In pandas, the column names were read from row 1 (0-indexed). So in openpyxl, headers should be row 1 or 2?
    # Let's write a quick loop to find headers.
    
# Actually, let's write the code to look at row 1 for headers.
headers = {sheet.cell(row=1, column=col).value: col for col in range(1, max_col + 1) if sheet.cell(row=1, column=col).value is not None}
print("Headers found in row 1:")
for h, c in headers.items():
    print(f"  {h} -> Col {c}")

# If row 1 is empty or doesn't have the election names, let's check row 2
if "Legislativas - 2025" not in headers:
    headers = {sheet.cell(row=2, column=col).value: col for col in range(1, max_col + 1) if sheet.cell(row=2, column=col).value is not None}
    print("Headers found in row 2 instead:")
    for h, c in headers.items():
         print(f"  {h} -> Col {c}")

# Find party name column and year column
party_col_name = "Unnamed: 44"  # Default name from pandas export
# Let's find which column has values like "PSD", "PS", etc.
party_col_idx = None
year_col_idx = None

for col in range(1, max_col + 1):
    for r in range(1, 5):
        val = sheet.cell(row=r, column=col).value
        if val == "PSD":
            party_col_idx = col
        if val == 2025:
            year_col_idx = col

# If not found by exact match, let's look at the mapping we exported
# In pandas, Unnamed: 44 was the clean party abbreviation.
# Let's search columns for a cell containing "PSD" in rows 2 to 30
if not party_col_idx:
    for col in range(1, max_col + 1):
        for row in range(2, 35):
            if sheet.cell(row=row, column=col).value == "PSD":
                party_col_idx = col
                break
        if party_col_idx:
            break

if not year_col_idx:
    for col in range(1, max_col + 1):
        for row in range(2, 35):
            if sheet.cell(row=row, column=col).value == 2025:
                year_col_idx = col
                break
        if year_col_idx:
            break

print(f"Party column index: {party_col_idx}")
print(f"Year column index: {year_col_idx}")

# Map party abbreviations in the sheet to row numbers
party_rows = {}
for row in range(2, max_row + 1):
    val = sheet.cell(row=row, column=party_col_idx).value if party_col_idx else None
    if val:
        party_rows[str(val).strip()] = row

# Map years to row numbers
year_rows = {}
for row in range(2, max_row + 1):
    val = sheet.cell(row=row, column=year_col_idx).value if year_col_idx else None
    if val:
        try:
            # cell value could be float like 2025.0
            yr = int(float(val))
            year_rows[yr] = row
        except:
            pass

print("Party rows:")
for p, r in party_rows.items():
    print(f"  {p} -> Row {r}")

print("Year rows:")
for y, r in year_rows.items():
    print(f"  {y} -> Row {r}")

# Reset all cells in the status matrix first, except for 'Extras' row
# The election columns start after party names and go until budgets.
election_cols = [c for h, c in headers.items() if any(x in str(h) for x in ["Legislativas", "Madeira", "Açores", "Europeias", "Autárquicas", "Presidenciais"])]
budget_cols = [c for h, c in headers.items() if "Orçamento" in str(h)]
principles_col = headers.get("Declaração de Princípios")

# Clean existing checkmarks (like 'Sim' or 'X') from party rows, but don't touch 'Extras' row which is at the end
# The Extras row has "Extras" in column 1 (Unnamed: 0)
extras_row = None
for r in range(2, max_row + 1):
    val = sheet.cell(row=r, column=1).value
    if val == "Extras":
        extras_row = r
        break

print(f"Extras row is: {extras_row}")

for r in range(2, max_row + 1):
    if r == extras_row:
        continue
    # Clear election cells
    for col in election_cols:
        sheet.cell(row=r, column=col).value = None
    # Clear budget cells
    for col in budget_cols:
        sheet.cell(row=r, column=col).value = None
    # Clear principles
    if principles_col:
        sheet.cell(row=r, column=principles_col).value = None

# Party mapping helper
def find_party_row(filename):
    lower = filename.lower()
    if "psd" in lower or "ppd" in lower:
        return party_rows.get("PSD")
    elif "ps" in lower and not "psd" in lower and not "psn" in lower:
        return party_rows.get("PS")
    elif "chega" in lower:
        return party_rows.get("CHEGA")
    elif "il" in lower or "iniciativa liberal" in lower:
        return party_rows.get("IL")
    elif "be" in lower or "bloco" in lower:
        return party_rows.get("BE")
    elif "cdu" in lower or "pcp" in lower or "pev" in lower:
        return party_rows.get("CDU - PCP/PEV")
    elif "livre" in lower:
        return party_rows.get("LIVRE")
    elif "pan" in lower:
        return party_rows.get("PAN")
    elif "cds" in lower:
        return party_rows.get("CDS")
    elif "adn" in lower or "pdr" in lower:
        return party_rows.get("ADN")
    elif "rir" in lower:
        return party_rows.get("R.I.R")
    elif "jpp" in lower:
        return party_rows.get("JPP")
    elif "nova direita" in lower or "nd" in lower or "nd " in lower:
        return party_rows.get("NOVA DIREITA")
    elif "pctp" in lower or "mrpp" in lower:
        return party_rows.get("PCTP/MRPP")
    elif "volt" in lower:
        return party_rows.get("VOLT PORTUGAL")
    elif "ergue-te" in lower or "pnr" in lower:
        return party_rows.get("ERGUE-TE/PNR")
    elif "mpt" in lower:
        return party_rows.get("MPT/ALTERNATIVA 21")
    elif "ptp" in lower:
        return party_rows.get("PTP")
    elif "nós" in lower or "nos cidadãos" in lower or "nós, cidadãos" in lower:
        return party_rows.get("NÓS, CIDADÃOS!")
    elif "ppm" in lower:
        return party_rows.get("PPM")
    elif "mas" in lower:
        return party_rows.get("MAS")
    elif "purp" in lower:
        return party_rows.get("PURP/(A)TUA")
    elif "mep" in lower:
        return party_rows.get("MEP")
    elif "pnd" in lower:
        return party_rows.get("PND")
    elif "ppv" in lower:
        return party_rows.get("PPV")
    elif "pous" in lower:
        return party_rows.get("POUS")
    elif "sda" in lower or "pda" in lower:
        return party_rows.get("PDA")
    elif "humanista" in lower or "ph" in lower:
        return party_rows.get("P.H.")
    elif "mms" in lower:
        return party_rows.get("MMS")
    elif "psn" in lower:
        return party_rows.get("PSN")
    elif "udp" in lower:
        return party_rows.get("UDP")
    elif "md" in lower:
        return party_rows.get("MD")
    elif "liberal social" in lower or "pls" in lower:
        return party_rows.get("Partido Liberal Social")
    elif "libertário" in lower or "pl" in lower:
        return party_rows.get("Partido Libertário")
    return None

# Write back original Extras
if extras_row:
    extras_mappings = {
        "Legislativas - 2025": "Manifesto chega 2019",
        "Madeira - 2025": "Manifesto IL",
        "Europeias - 2024": "Constituição",
        "Madeira - 2024": "Programa Cotrim - 19/21",
        "Legislativas - 2024": "PREC Liberal",
        "Açores - 2024": "Programa Cotrim - 21/23",
        "Madeira - 2023": "Programa Rui Rocha - 23/25"
    }
    for col_name, value in extras_mappings.items():
        col_idx = headers.get(col_name)
        if col_idx:
            sheet.cell(row=extras_row, column=col_idx).value = value


# Scan and map files
unmapped_files = []
mapped_count = 0

for root, dirs, files in os.walk(data_dir):
    for file in files:
        if not file.endswith(".pdf"):
            continue
        
        filepath = os.path.join(root, file)
        rel_path = os.path.relpath(filepath, data_dir)
        parts = rel_path.split(os.sep)
        category = parts[0]
        
        # Get year from file name or directory
        year_match = re.search(r"\b(19\d{2}|20\d{2})\b", file)
        if not year_match:
            # check directory name
            year_match = re.search(r"\b(19\d{2}|20\d{2})\b", root)
        
        year = int(year_match.group(1)) if year_match else None
        
        # Special case for Chega 70 measures
        if "70-medidas-para-reerguer-portugal-chega" in file.lower():
            year = 2024
            category = "Legislativas"
            
        # 1. State Budget mapping
        if category == "Orçamentos de Estado":
            if year and year in year_rows:
                r_idx = year_rows[year]
                # Determine budget type
                lower_file = file.lower()
                col_name = "Orçamento de Estado"
                if "madeira" in lower_file:
                    col_name = "Orçamento - Madeira"
                elif "açores" in lower_file:
                    col_name = "Orçamento - Açores"
                elif "ue" in lower_file or "europeu" in lower_file:
                    col_name = "Orçamento da UE"
                
                col_idx = headers.get(col_name)
                if col_idx:
                    sheet.cell(row=r_idx, column=col_idx).value = "Sim"
                    mapped_count += 1
                else:
                    unmapped_files.append((rel_path, "Budget column not found"))
            else:
                unmapped_files.append((rel_path, "Year row not found"))
            continue
            
        # 2. Declaração de Princípios mapping
        if category == "Declaração de Princípios":
            r_idx = find_party_row(file)
            col_idx = headers.get("Declaração de Princípios")
            if r_idx and col_idx:
                sheet.cell(row=r_idx, column=col_idx).value = "Sim"
                mapped_count += 1
            else:
                unmapped_files.append((rel_path, "Party or principles column not found"))
            continue

        # 3. Constitution mapping
        if file.lower() == "constituição.pdf":
            # The constitution is listed in Extras under Europeias - 2024 in the original file
            # Let's keep it as is, or put it in Extras row under a Constitution column if we want,
            # or just leave it. Let's see: Extras has it already.
            continue
            
        # 4. Other root files
        if root == data_dir and file.lower() != "constituição.pdf":
            # E.g. 70-MEDIDAS-PARA-REERGUER-PORTUGAL-CHEGA.pdf (already set category to Legislativas and year to 2024 above)
            r_idx = find_party_row(file)
            if r_idx and year:
                col_name = f"Legislativas - {year}"
                col_idx = headers.get(col_name)
                if col_idx:
                    sheet.cell(row=r_idx, column=col_idx).value = "Sim"
                    mapped_count += 1
                else:
                    unmapped_files.append((rel_path, f"Column {col_name} not found"))
            else:
                unmapped_files.append((rel_path, "Party or year not identified"))
            continue
            
        # 5. Elections mapping (Legislativas, Açores, Madeira, Europeias)
        if year:
            # Map coalition AD to both PSD and CDS
            is_ad = "ad " in file.lower() or " ad_" in file.lower() or "ad_açores" in file.lower() or "ad_madeira" in file.lower() or "aliança democrática" in file.lower()
            
            parties_to_map = []
            if is_ad:
                parties_to_map = [party_rows.get("PSD"), party_rows.get("CDS")]
            else:
                r_idx = find_party_row(file)
                if r_idx:
                    parties_to_map = [r_idx]
            
            col_name = f"{category} - {year}" if category != "Europeias" else f"Europeias - {year}"
            # Special case for column naming:
            # Europeias in spreadsheet is "Europeias - 2024", "Europeias - 2019", "Europeias - 2009", "Europeias - 2004", "Europeias 1999" (without dash)
            if category == "Europeias" and year == 1999:
                col_name = "Europeias 1999"
                
            col_idx = headers.get(col_name)
            
            if col_idx and parties_to_map:
                for r_idx in parties_to_map:
                    if r_idx:
                        # Write "Sim" or the filename if we want. Let's write "Sim"
                        sheet.cell(row=r_idx, column=col_idx).value = "Sim"
                        mapped_count += 1
            else:
                unmapped_files.append((rel_path, f"Column '{col_name}' or Party not found"))
        else:
            unmapped_files.append((rel_path, "Year not found"))

# Save workbook
wb.save(excel_path)
print(f"\nDone! Mapped {mapped_count} files successfully.")
if unmapped_files:
    print(f"\nThere were {len(unmapped_files)} unmapped files:")
    for f, reason in unmapped_files[:20]:
        print(f"  - {f}: {reason}")
    if len(unmapped_files) > 20:
        print(f"  - ... and {len(unmapped_files) - 20} more.")
