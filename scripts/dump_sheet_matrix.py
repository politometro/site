import openpyxl
import json
import os

wb = openpyxl.load_workbook("Polítólogo Português.xlsx", data_only=True)
sheet = wb.active

# Parse headers
headers = []
for col_idx in range(1, sheet.max_column + 1):
    val = sheet.cell(row=1, column=col_idx).value
    headers.append(val if val is not None else f"Unnamed_{col_idx}")

print("Clean Headers:", headers)

matrix = []

# Helper to safely get hex color from cell fill
def get_cell_color(cell):
    fill = cell.fill
    if not fill or fill.fill_type != "solid" or not fill.start_color:
        return None
    try:
        color_val = fill.start_color
        if hasattr(color_val, "rgb"):
            return str(color_val.rgb)
        return str(color_val)
    except Exception as e:
        return "UNKNOWN_COLOR_ERROR"

for row_idx in range(2, sheet.max_row + 1):
    row_cells = [sheet.cell(row=row_idx, column=c) for c in range(1, sheet.max_column + 1)]
    party_val = row_cells[0].value
    if party_val is None:
        continue
        
    row_data = {
        "party": party_val,
        "cells": []
    }
    
    for col_idx, cell in enumerate(row_cells):
        col_name = headers[col_idx]
        val = cell.value
        
        # Safe color extraction
        color = None
        fill = cell.fill
        if fill and fill.fill_type == "solid" and fill.start_color:
            try:
                color = str(fill.start_color.rgb)
            except:
                try:
                    color = str(fill.start_color)
                except:
                    color = "ERROR"
                    
        row_data["cells"].append({
            "col": col_name,
            "col_idx": col_idx,
            "value": val,
            "color": color
        })
        
    matrix.append(row_data)

print(f"Parsed {len(matrix)} rows of party data.")

# Write to a file for analysis
with open("scripts/excel_parsed.json", "w", encoding="utf-8") as f:
    json.dump(matrix, f, indent=2, ensure_ascii=False)

print("Saved parsing to scripts/excel_parsed.json")
