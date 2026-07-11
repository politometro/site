import openpyxl

wb = openpyxl.load_workbook("Polítólogo Português.xlsx", data_only=True)
sheet = wb.active

headers = [cell.value for cell in sheet[1]]
print("Headers length:", len(headers))
print("Headers:", headers)

unique_colors = set()
for r in range(1, sheet.max_row + 1):
    for c in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=r, column=c)
        fill = cell.fill
        if fill and fill.fill_type == "solid" and fill.start_color:
            unique_colors.add(fill.start_color.rgb)

print("Unique solid fill colors in sheet:", unique_colors)
